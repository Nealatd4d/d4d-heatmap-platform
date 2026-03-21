#!/usr/bin/env python3
"""
Lake County IL → Supabase Postgres loader.

Parses SOE Software HTML-XML .xls election files and loads:
  - jurisdictions (lake_county)
  - precincts (from GeoJSON TWP_PCT + election data)
  - elections (mapped from file dates/types)
  - districts, races, candidates, results, turnout

Precinct ID formula matches existing convention: MD5(jurisdiction_id + ':' + source_precinct_name)
but uses make_id(jurisdiction_id, source_precinct_name) → MD5(lower('lake_county|source_name'))[:16]

Election file mapping:
  detailxls_2:  2021-03-12 → Municipal  (North Chicago/Waukegan)
  detailxls_3:  2021-04-22 → Consolidated
  detailxls_4:  2022-07-14 → Primary (2022 Primary)
  detailxls_5:  2022-11-30 → General (2022 General)
  detailxls_6:  2023-03-16 → Municipal
  detailxls_7:  2023-04-20 → Consolidated
  detailxls_8:  2024-11-22 → General (2024 General)
  detailxls_9:  2025-03-14 → Municipal
  detailxls_10: 2025-04-21 → Consolidated
  detailxlsx_2: Lake County consolidated (Excel format - separate handling)
"""

import hashlib
import io
import json
import os
import re
import sys
import time
import xml.etree.ElementTree as ET

import psycopg2

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from classify_race import classify_race_type
from normalize_race import normalize_race_name

# ── Connection ──────────────────────────────────────────────
DB_URI = os.environ.get(
    "DATABASE_URL",
    "postgresql://postgres.nfjfqdffulhqhszhlymo:D4dHeatmap2026!Pg"
    "@aws-0-us-west-2.pooler.supabase.com:5432/postgres"
)

JURISDICTION_ID = "lake_county"
COUNTY_DIR = os.environ.get("COUNTY_DIR", "/home/user/workspace/county_expansion")

NS = {
    's': 'urn:schemas-microsoft-com:office:spreadsheet',
    'o': 'urn:schemas-microsoft-com:office:office',
    'x': 'urn:schemas-microsoft-com:office:excel'
}

# ── Election file → election metadata mapping ──────────────
# Each tuple: (folder_name, election_id, election_name, date, type, year)
ELECTION_FILES = [
    ("detailxls_2",  "lake_2021_municipal_mar", "2021 Lake County Municipal (March)", "2021-03-09", "municipal", 2021),
    ("detailxls_3",  "lake_2021_consolidated",  "2021 Lake County Consolidated",       "2021-04-06", "consolidated", 2021),
    ("detailxls_4",  "2022_primary",            "2022 Primary Election",               "2022-06-28", "primary", 2022),
    ("detailxls_5",  "2022_general",            "2022 General Election",               "2022-11-08", "general", 2022),
    ("detailxls_6",  "lake_2023_municipal_mar", "2023 Lake County Municipal (March)", "2023-02-28", "municipal", 2023),
    ("detailxls_7",  "lake_2023_consolidated",  "2023 Lake County Consolidated",       "2023-04-04", "consolidated", 2023),
    ("detailxls_8",  "2024_general",            "2024 General Election",               "2024-11-05", "general", 2024),
    ("detailxls_9",  "lake_2025_municipal_mar", "2025 Lake County Municipal (March)", "2025-02-25", "municipal", 2025),
    ("detailxls_10", "lake_2025_consolidated",  "2025 Lake County Consolidated",       "2025-04-01", "consolidated", 2025),
]


# ── Helpers ─────────────────────────────────────────────────
def make_id(*parts):
    """Create a deterministic 16-char hex ID from parts."""
    raw = "|".join(str(p).strip().lower() for p in parts)
    return hashlib.md5(raw.encode()).hexdigest()[:16]


def safe_int(val):
    if val is None or val == '':
        return 0
    try:
        return int(float(str(val).replace(',', '').replace('%', '').strip()))
    except (ValueError, TypeError):
        return 0


def safe_float(val):
    if val is None or val == '':
        return 0.0
    try:
        return float(str(val).replace(',', '').replace('%', '').strip())
    except (ValueError, TypeError):
        return 0.0


def parse_lake_precinct(name):
    """Parse Lake County precinct name. Format: 'Township NN' e.g. 'Antioch 1', 'West Deerfield 415'.
    Returns (township_name, precinct_number). Ward column is integer in DB,
    so we store township info in source_precinct_name only."""
    name = name.strip()
    m = re.match(r'^(.+?)\s+(\d+)$', name)
    if m:
        precinct_num = int(m.group(2))
        return None, precinct_num  # ward=None for Lake County (townships, not wards)
    return None, None


def classify_lake_contest(contest_name):
    """Classify a Lake County contest name into (district_type_id, district_number).
    Uses the unified classifier first, then handles Lake-County-specific contests."""
    c = contest_name.strip()
    if not c:
        return ('referendum', 0)
    
    cu = c.upper()
    
    # Strip "(Vote For N)" suffix for classification
    cleaned = re.sub(r'\s*\(Vote\s+For\s+\d+\)\s*$', '', c, flags=re.IGNORECASE).strip()
    cleaned_u = cleaned.upper()
    
    # Federal/state-wide races
    if 'PRESIDENT' in cleaned_u and ('UNITED STATES' in cleaned_u or 'VICE' in cleaned_u or cleaned_u.startswith('PRESIDENT')):
        return ('president', 0)
    if 'U.S. SENATOR' in cleaned_u or 'UNITED STATES SENATOR' in cleaned_u:
        return ('us_senate', 0)
    if 'GOVERNOR' in cleaned_u and 'LT' not in cleaned_u and 'LIEUTENANT' not in cleaned_u:
        return ('governor', 0)
    if 'LIEUTENANT GOVERNOR' in cleaned_u or 'LT. GOVERNOR' in cleaned_u:
        return ('lt_governor', 0)
    if 'ATTORNEY GENERAL' in cleaned_u:
        return ('attorney_general', 0)
    if 'SECRETARY OF STATE' in cleaned_u:
        return ('secretary_of_state', 0)
    if 'COMPTROLLER' in cleaned_u:
        return ('comptroller', 0)
    if 'TREASURER' in cleaned_u and 'CITY' not in cleaned_u and 'VILLAGE' not in cleaned_u and 'TOWNSHIP' not in cleaned_u:
        return ('treasurer_state', 0)
    
    # Congressional
    m = re.search(r'(\d+)\w*\s*(TH|ST|ND|RD)?\s*CONGRESS', cleaned_u)
    if m:
        return ('congress', int(m.group(1)))
    m = re.search(r'U\.?S\.?\s+REPRESENTATIVE.*?(\d+)', cleaned_u)
    if m:
        return ('congress', int(m.group(1)))
    
    # State legislature
    m = re.search(r'(\d+)\w*\s*(TH|ST|ND|RD)?\s*(STATE\s+)?SENATE', cleaned_u)
    if m:
        return ('state_senate', int(m.group(1)))
    m = re.search(r'STATE\s+SENATOR.*?(\d+)', cleaned_u)
    if m:
        return ('state_senate', int(m.group(1)))
    m = re.search(r'(\d+)\w*\s*(TH|ST|ND|RD)?\s*(STATE\s+)?REPRESENTATIVE', cleaned_u)
    if m:
        return ('state_rep', int(m.group(1)))
    m = re.search(r'STATE\s+REPRESENTATIVE.*?(\d+)', cleaned_u)
    if m:
        return ('state_rep', int(m.group(1)))
    
    # Judicial
    if 'SUPREME' in cleaned_u:
        m = re.search(r'(\d+)', cleaned_u)
        return ('supreme_court', int(m.group(1)) if m else 1)
    if 'APPELLATE' in cleaned_u:
        m = re.search(r'(\d+)', cleaned_u)
        return ('appellate_court', int(m.group(1)) if m else 2)
    if 'SUBCIRCUIT' in cleaned_u or 'SUB-CIRCUIT' in cleaned_u or 'SUB CIRCUIT' in cleaned_u:
        m = re.search(r'(\d+)', cleaned_u)
        return ('subcircuit', int(m.group(1)) if m else 0)
    if 'CIRCUIT' in cleaned_u and ('JUDGE' in cleaned_u or 'COURT' in cleaned_u or 'JUDICIAL' in cleaned_u):
        m = re.search(r'(\d+)', cleaned_u)
        return ('circuit_court', int(m.group(1)) if m else 0)
    
    # Lake County-specific county offices
    if 'COUNTY BOARD' in cleaned_u:
        m = re.search(r'DISTRICT\s+(\d+)', cleaned_u)
        return ('lake_county_board', int(m.group(1)) if m else 0)
    if 'COUNTY CLERK' in cleaned_u:
        return ('lake_county_clerk', 0)
    if 'COUNTY CORONER' in cleaned_u:
        return ('lake_county_coroner', 0)
    if 'COUNTY AUDITOR' in cleaned_u:
        return ('lake_county_auditor', 0)
    if 'COUNTY RECORDER' in cleaned_u:
        return ('lake_county_recorder', 0)
    if 'COUNTY SHERIFF' in cleaned_u:
        return ('lake_county_sheriff', 0)
    if 'COUNTY TREASURER' in cleaned_u:
        return ('lake_county_treasurer', 0)
    if 'COUNTY STATE' in cleaned_u and 'ATTORNEY' in cleaned_u:
        return ('lake_county_sa', 0)
    if 'FOREST PRESERVE' in cleaned_u:
        m = re.search(r'DISTRICT\s+(\d+)', cleaned_u)
        return ('lake_forest_preserve', int(m.group(1)) if m else 0)
    if 'REGIONAL SUPERINTENDENT' in cleaned_u or 'REGIONAL BOARD' in cleaned_u:
        return ('regional_supt', 0)
    
    # Township offices - use allowed race_type 'township_office'
    if 'TOWNSHIP' in cleaned_u and ('SUPERVISOR' in cleaned_u or 'ASSESSOR' in cleaned_u or
                                     'CLERK' in cleaned_u or 'TRUSTEE' in cleaned_u or
                                     'HIGHWAY' in cleaned_u or 'HWY' in cleaned_u or
                                     'COLLECTOR' in cleaned_u or 'COMMISSIONER' in cleaned_u):
        return ('township', 0)
    
    # Municipal - city/village offices
    if any(prefix in cleaned_u for prefix in ['CITY OF', 'VILLAGE OF', 'TOWN OF']):
        return ('municipal', 0)
    
    # Specific special district types (each has its own allowed race_type)
    if 'LIBRARY' in cleaned_u:
        return ('library', 0)
    if 'PARK' in cleaned_u:
        return ('park_district', 0)
    if 'FIRE' in cleaned_u:
        return ('fire_district', 0)
    if any(kw in cleaned_u for kw in ['SCHOOL', 'EDUCATION', 'COMMUNITY COLLEGE']):
        return ('school_district', 0)
    if 'WATER' in cleaned_u or 'SANITARY' in cleaned_u:
        return ('water_reclamation', 0)
    
    # State central committee / delegates
    if 'STATE CENTRAL' in cleaned_u or 'COMMITTEEMAN' in cleaned_u or 'COMMITTEEWOMAN' in cleaned_u:
        m = re.search(r'(\d+)', cleaned_u)
        return ('state_central_comm', int(m.group(1)) if m else 0)
    if 'DELEGATE' in cleaned_u:
        m = re.search(r'(\d+)', cleaned_u)
        return ('delegate', int(m.group(1)) if m else 0)
    if 'PRECINCT COMMITTEE' in cleaned_u:
        return ('precinct_committee', 0)
    
    # Referendum / ballot question
    if any(kw in cleaned_u for kw in ['REFERENDUM', 'PROPOSITION', 'QUESTION', 'ADVISORY',
                                       'SHALL', 'BINDING']):
        return ('referendum', 0)
    
    # Catch-all
    return ('local_other', 0)


# ── Map district_type_id to valid race_type ──────────────────
DISTRICT_TO_RACE_TYPE = {
    'president': 'president',
    'us_senate': 'us_senate',
    'congress': 'congress',
    'governor': 'governor',
    'lt_governor': 'lt_governor',
    'attorney_general': 'attorney_general',
    'secretary_of_state': 'secretary_of_state',
    'comptroller': 'comptroller',
    'treasurer_state': 'treasurer_state',
    'state_senate': 'state_senate',
    'state_rep': 'state_rep',
    'state_central_comm': 'state_central_comm',
    'delegate': 'delegate',
    'supreme_court': 'supreme_court',
    'appellate_court': 'appellate_court',
    'circuit_court': 'circuit_court',
    'subcircuit': 'subcircuit',
    'referendum': 'referendum',
    'municipal': 'municipal',
    'township': 'township_office',
    'library': 'library',
    'park_district': 'park_district',
    'fire_district': 'fire_district',
    'school_district': 'school_district',
    'water_reclamation': 'water_reclamation',
    'precinct_committee': 'township_committeeperson',
    # Lake County-specific → map to closest allowed race_type
    'lake_county_board': 'other',
    'lake_county_clerk': 'other',
    'lake_county_coroner': 'other',
    'lake_county_auditor': 'other',
    'lake_county_recorder': 'other',
    'lake_county_sheriff': 'other',
    'lake_county_treasurer': 'other',
    'lake_county_sa': 'other',
    'lake_forest_preserve': 'other',
    'regional_supt': 'other',
    'local_other': 'other',
}


def get_race_type(district_type_id):
    """Map a district_type_id to a valid races.race_type value."""
    return DISTRICT_TO_RACE_TYPE.get(district_type_id, 'other')


# ── XML Parsing ─────────────────────────────────────────────
def get_cell_values(row):
    """Extract all cell values from an XML Row, handling MergeAcross."""
    values = []
    for cell in row.findall('s:Cell', NS):
        # Handle Index attribute (skipped cells)
        idx_attr = cell.get('{urn:schemas-microsoft-com:office:spreadsheet}Index')
        if idx_attr:
            target_idx = int(idx_attr) - 1  # 1-based to 0-based
            while len(values) < target_idx:
                values.append('')
        
        data = cell.find('s:Data', NS)
        val = data.text if data is not None and data.text else ''
        values.append(val)
        
        # Handle MergeAcross (merged cells span additional columns)
        merge = cell.get('{urn:schemas-microsoft-com:office:spreadsheet}MergeAcross')
        if merge:
            for _ in range(int(merge)):
                values.append('')
    
    return values


def parse_xml_election_file(filepath, election_id, source_label):
    """Parse a SOE Software HTML-XML .xls file.
    
    Structure per contest worksheet:
      Row 0: [contest_name]  (merged cell)
      Row 1: ['', '', candidate1, candidate2, ...]  (candidate names between the per-candidate column groups)
      Row 2: ['County', 'Registered Voters', 'ED', 'EV', 'VBM', 'Late', 'Total Votes', ...repeated per candidate..., 'Total']
      Row 3+: [precinct_name, reg_voters, ed1, ev1, vbm1, late1, total1, ed2, ev2, vbm2, late2, total2, ..., grand_total]
      Last row: ['Total:', ...]
    
    The 'Registered Voters' worksheet has turnout data:
      Row 0: ['County', 'Registered Voters', 'Election Day', 'Early Voting', 'Vote By Mail (VBM)', 'Late Arriving VBM/Prov', 'Ballots Cast', 'Voter Turnout']
      Row 1+: [precinct_name, reg, ed, ev, vbm, late, ballots, turnout_pct]
    """
    precincts = {}
    districts = {}
    races = {}
    candidates = {}
    results = []
    turnout = []
    pd_set = set()
    
    tree = ET.parse(filepath)
    root = tree.getroot()
    
    total_result_rows = 0
    
    for ws in root.findall('.//s:Worksheet', NS):
        ws_name = ws.get('{urn:schemas-microsoft-com:office:spreadsheet}Name', 'unnamed')
        
        table = ws.find('s:Table', NS)
        if table is None:
            continue
        
        rows = table.findall('s:Row', NS)
        if len(rows) < 2:
            continue
        
        # ── Handle "Registered Voters" worksheet (turnout) ──
        if ws_name == 'Registered Voters':
            for row in rows[1:]:  # skip header
                vals = get_cell_values(row)
                if not vals or not vals[0] or vals[0] == 'Total:':
                    continue
                
                pname = vals[0].strip()
                if not pname:
                    continue
                
                pid = make_id(JURISDICTION_ID, pname)
                township, pnum = parse_lake_precinct(pname)
                
                if pid not in precincts:
                    precincts[pid] = (pid, JURISDICTION_ID, pname, township, pnum)
                
                reg = safe_int(vals[1]) if len(vals) > 1 else 0
                ballots = safe_int(vals[6]) if len(vals) > 6 else 0
                turnout_pct_str = vals[7] if len(vals) > 7 else '0'
                turnout_pct = safe_float(turnout_pct_str)
                
                if reg > 0:
                    turnout.append((election_id, pid, reg, ballots, turnout_pct, source_label, pname))
            
            continue
        
        # ── Handle "Table of Contents" worksheet ──
        if ws_name == 'Table of Contents':
            continue
        
        # ── Handle contest worksheets ──
        # Row 0: contest name
        row0_vals = get_cell_values(rows[0])
        contest_name = row0_vals[0].strip() if row0_vals else ''
        if not contest_name:
            continue
        
        # Row 1: candidate names (positioned at column offsets)
        # The pattern: columns are grouped per candidate. Each group has 5 data cols 
        # (ED, EV, VBM, Late, Total). Plus 2 header cols (County, Registered Voters).
        # So candidates are at columns: 2, 7, 12, 17, ...
        # But in Row 1, the candidate names appear at those positions.
        row1_vals = get_cell_values(rows[1]) if len(rows) > 1 else []
        
        # Row 2: full column headers - this tells us the exact structure
        row2_vals = get_cell_values(rows[2]) if len(rows) > 2 else []
        
        # Count how many "Total Votes" appear to determine candidate count
        # Each candidate block has: ED, EV, VBM, Late, Total Votes (5 cols)
        # Header cols: County, Registered Voters (2 cols)
        # Last column: Total (grand total)
        
        # Find candidate names from row 1
        cand_names = []
        for i, v in enumerate(row1_vals):
            v = v.strip()
            if v and v not in ('', 'County', 'Registered Voters', 'Total', 'Election Day',
                                'Early Voting', 'Vote By Mail (VBM)', 'Late Arriving VBM/Prov',
                                'Total Votes'):
                cand_names.append((i, v))
        
        if not cand_names:
            # Sometimes single candidate or no candidates parsed
            continue
        
        # Determine per-candidate column block size from header row
        # Find positions of "Total Votes" in header (row 2)
        total_votes_positions = []
        for i, v in enumerate(row2_vals):
            if 'Total Votes' in str(v):
                total_votes_positions.append(i)
        
        if not total_votes_positions:
            # Try finding "Votes" columns
            for i, v in enumerate(row2_vals):
                if v.strip() == 'Votes':
                    total_votes_positions.append(i)
        
        if not total_votes_positions:
            continue
        
        # Number of candidates should match number of "Total Votes" columns
        n_cands = len(total_votes_positions)
        
        # Sometimes there's a mismatch — use whichever is smaller
        if len(cand_names) > n_cands:
            cand_names = cand_names[:n_cands]
        elif len(cand_names) < n_cands:
            n_cands = len(cand_names)
        
        # Classify the contest
        dt_id, dt_num = classify_lake_contest(contest_name)
        did = make_id(dt_id, dt_num)
        if did not in districts:
            dname = f"{dt_id.replace('_', ' ').title()} {dt_num}" if dt_num else dt_id.replace('_', ' ').title()
            districts[did] = (did, dt_id, dt_num, dname)
        
        # Normalize race name for cross-county consistency
        normalized_name = normalize_race_name(contest_name)

        # Create race
        rid = make_id(election_id, normalized_name)
        if rid not in races:
            races[rid] = (rid, election_id, did, normalized_name, contest_name, get_race_type(dt_id))
        
        # Create candidate entries
        cand_ids = []
        for _, cname in cand_names:
            cid = make_id(cname, '')  # No party info in these files
            if cid not in candidates:
                candidates[cid] = (cid, cname, None)
            cand_ids.append(cid)
        
        # Parse data rows (row 3 onwards until 'Total:')
        for row in rows[3:]:
            vals = get_cell_values(row)
            if not vals or not vals[0] or vals[0].strip() == 'Total:':
                continue
            
            pname = vals[0].strip()
            if not pname:
                continue
            
            pid = make_id(JURISDICTION_ID, pname)
            township, pnum = parse_lake_precinct(pname)
            
            if pid not in precincts:
                precincts[pid] = (pid, JURISDICTION_ID, pname, township, pnum)
            
            # Track precinct-district mapping
            pd_key = (pid, did)
            if pd_key not in pd_set:
                pd_set.add(pd_key)
            
            # Extract votes per candidate using Total Votes column positions
            for c_idx in range(min(n_cands, len(cand_ids))):
                tv_col = total_votes_positions[c_idx]
                votes = safe_int(vals[tv_col]) if tv_col < len(vals) else 0
                
                results.append((
                    election_id, pid, rid, cand_ids[c_idx], votes,
                    source_label, contest_name, pname, cand_names[c_idx][1]
                ))
                total_result_rows += 1
    
    print(f"  {source_label}: {total_result_rows:,} results, {len(turnout):,} turnout, "
          f"{len(races)} races, {len(candidates)} candidates, {len(precincts)} precincts")
    return precincts, districts, races, candidates, results, turnout, pd_set


def parse_xlsx_election_file(filepath, election_id, source_label):
    """Parse a true .xlsx file using openpyxl."""
    try:
        import openpyxl
    except ImportError:
        print(f"  SKIPPING {source_label}: openpyxl not installed")
        return {}, {}, {}, {}, [], [], set()
    
    precincts = {}
    districts = {}
    races = {}
    candidates = {}
    results = []
    turnout = []
    pd_set = set()
    
    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    total_result_rows = 0
    
    for ws_name in wb.sheetnames:
        ws = wb[ws_name]
        rows_data = []
        for row in ws.iter_rows(values_only=True):
            rows_data.append([str(v) if v is not None else '' for v in row])
        
        if len(rows_data) < 3:
            continue
        
        # Similar logic to XML parser
        if ws_name == 'Registered Voters' or 'registered' in ws_name.lower():
            for vals in rows_data[1:]:
                if not vals or not vals[0] or vals[0] == 'Total:':
                    continue
                pname = vals[0].strip()
                if not pname:
                    continue
                pid = make_id(JURISDICTION_ID, pname)
                township, pnum = parse_lake_precinct(pname)
                if pid not in precincts:
                    precincts[pid] = (pid, JURISDICTION_ID, pname, township, pnum)
                reg = safe_int(vals[1]) if len(vals) > 1 else 0
                ballots = safe_int(vals[6]) if len(vals) > 6 else 0
                turnout_pct = safe_float(vals[7]) if len(vals) > 7 else 0.0
                if reg > 0:
                    turnout.append((election_id, pid, reg, ballots, turnout_pct, source_label, pname))
            continue
        
        if ws_name == 'Table of Contents' or 'contents' in ws_name.lower():
            continue
        
        # Contest worksheet
        contest_name = rows_data[0][0].strip() if rows_data[0] else ''
        if not contest_name:
            continue
        
        row1 = rows_data[1] if len(rows_data) > 1 else []
        row2 = rows_data[2] if len(rows_data) > 2 else []
        
        cand_names = []
        for i, v in enumerate(row1):
            v = v.strip()
            if v and v not in ('', 'County', 'Registered Voters', 'Total', 'Election Day',
                                'Early Voting', 'Vote By Mail (VBM)', 'Late Arriving VBM/Prov',
                                'Total Votes', 'None'):
                cand_names.append((i, v))
        
        total_votes_positions = []
        for i, v in enumerate(row2):
            if 'Total Votes' in str(v):
                total_votes_positions.append(i)
        
        if not cand_names or not total_votes_positions:
            continue
        
        n_cands = min(len(cand_names), len(total_votes_positions))
        cand_names = cand_names[:n_cands]
        
        dt_id, dt_num = classify_lake_contest(contest_name)
        did = make_id(dt_id, dt_num)
        if did not in districts:
            dname = f"{dt_id.replace('_', ' ').title()} {dt_num}" if dt_num else dt_id.replace('_', ' ').title()
            districts[did] = (did, dt_id, dt_num, dname)
        
        normalized_name = normalize_race_name(contest_name)

        rid = make_id(election_id, normalized_name)
        if rid not in races:
            races[rid] = (rid, election_id, did, normalized_name, contest_name, get_race_type(dt_id))

        cand_ids = []
        for _, cname in cand_names:
            cid = make_id(cname, '')
            if cid not in candidates:
                candidates[cid] = (cid, cname, None)
            cand_ids.append(cid)

        for vals in rows_data[3:]:
            if not vals or not vals[0] or vals[0].strip() == 'Total:':
                continue
            pname = vals[0].strip()
            if not pname:
                continue
            pid = make_id(JURISDICTION_ID, pname)
            township, pnum = parse_lake_precinct(pname)
            if pid not in precincts:
                precincts[pid] = (pid, JURISDICTION_ID, pname, township, pnum)
            pd_key = (pid, did)
            if pd_key not in pd_set:
                pd_set.add(pd_key)
            for c_idx in range(min(n_cands, len(cand_ids))):
                tv_col = total_votes_positions[c_idx]
                votes = safe_int(vals[tv_col]) if tv_col < len(vals) else 0
                results.append((
                    election_id, pid, rid, cand_ids[c_idx], votes,
                    source_label, contest_name, pname, cand_names[c_idx][1]
                ))
                total_result_rows += 1
    
    wb.close()
    print(f"  {source_label}: {total_result_rows:,} results, {len(turnout):,} turnout, "
          f"{len(races)} races, {len(candidates)} candidates, {len(precincts)} precincts")
    return precincts, districts, races, candidates, results, turnout, pd_set


# ── Bulk copy helper ────────────────────────────────────────
def copy_tuples(cur, table, columns, rows):
    if not rows:
        return 0
    buf = io.StringIO()
    for row in rows:
        line = '\t'.join('' if v is None else str(v) for v in row)
        buf.write(line + '\n')
    buf.seek(0)
    cols = ', '.join(columns)
    cur.copy_expert(f"COPY {table} ({cols}) FROM STDIN WITH (FORMAT text, NULL '')", buf)
    return len(rows)


# ── Main ────────────────────────────────────────────────────
def main():
    data_dir = f"{COUNTY_DIR}/election_data"
    
    print("=" * 60)
    print("LAKE COUNTY ELECTION DATA LOADER")
    print("=" * 60)
    t0 = time.time()
    
    # ── Phase 1: Scan all election files ──
    all_p, all_d, all_r, all_c = {}, {}, {}, {}
    all_results, all_turnout = [], []
    all_pd = set()
    election_meta = []  # (id, name, date, type, year)
    
    for folder, eid, ename, edate, etype, eyear in ELECTION_FILES:
        filepath = f"{data_dir}/{folder}/detail.xls"
        if not os.path.exists(filepath):
            print(f"  MISSING: {filepath}")
            continue
        
        source_label = f"Lake County {ename}"
        p, d, r, c, res, to, pd = parse_xml_election_file(filepath, eid, source_label)
        all_p.update(p)
        all_d.update(d)
        all_r.update(r)
        all_c.update(c)
        all_results.extend(res)
        all_turnout.extend(to)
        all_pd.update(pd)
        election_meta.append((eid, ename, edate, etype, eyear))
    
    # Also try the xlsx file
    xlsx_path = f"{data_dir}/detailxlsx_2/detail.xlsx"
    if os.path.exists(xlsx_path):
        eid = "lake_consolidated_xlsx"
        ename = "Lake County Consolidated (XLSX)"
        source_label = f"Lake County {ename}"
        p, d, r, c, res, to, pd = parse_xlsx_election_file(
            xlsx_path, eid, source_label
        )
        if res:
            all_p.update(p)
            all_d.update(d)
            all_r.update(r)
            all_c.update(c)
            all_results.extend(res)
            all_turnout.extend(to)
            all_pd.update(pd)
            # Try to determine the election date from the file
            election_meta.append((eid, ename, "2025-04-01", "consolidated", 2025))
    
    scan_time = time.time() - t0
    print(f"\nScan complete in {scan_time:.1f}s")
    print(f"  Precincts:  {len(all_p):,}")
    print(f"  Districts:  {len(all_d):,}")
    print(f"  Races:      {len(all_r):,}")
    print(f"  Candidates: {len(all_c):,}")
    print(f"  Results:    {len(all_results):,}")
    print(f"  Turnout:    {len(all_turnout):,}")
    print(f"  PD maps:    {len(all_pd):,}")
    print(f"  Elections:  {len(election_meta)}")
    
    if not all_results:
        print("\nERROR: No results parsed!")
        sys.exit(1)
    
    # ── Phase 2: Load to Supabase ──
    print(f"\nConnecting to Supabase Postgres...")
    conn = psycopg2.connect(DB_URI, connect_timeout=30)
    conn.autocommit = False
    cur = conn.cursor()
    cur.execute("SET statement_timeout = 0;")
    
    # ── Create new district types needed for Lake County ──
    print(f"\n--- District Types ---")
    # (id, name, category)
    new_district_types = [
        ('municipal', 'Municipal Office', 'municipal'),
        ('township', 'Township Office', 'municipal'),
        ('park_district', 'Park District', 'special'),
        ('library', 'Library District', 'special'),
        ('fire_district', 'Fire Protection District', 'special'),
        ('school_district', 'School District', 'school'),
        ('water_reclamation', 'Water/Sanitary District', 'special'),
        ('lake_county_board', 'Lake County Board District', 'county'),
        ('lake_county_clerk', 'Lake County Clerk', 'county'),
        ('lake_county_coroner', 'Lake County Coroner', 'county'),
        ('lake_county_auditor', 'Lake County Auditor', 'county'),
        ('lake_county_recorder', 'Lake County Recorder', 'county'),
        ('lake_county_sheriff', 'Lake County Sheriff', 'county'),
        ('lake_county_treasurer', 'Lake County Treasurer', 'county'),
        ('lake_county_sa', 'Lake County State\'s Attorney', 'county'),
        ('lake_forest_preserve', 'Lake County Forest Preserve District', 'special'),
        ('regional_supt', 'Regional Superintendent of Schools', 'school'),
        ('precinct_committee', 'Precinct Committeeperson', 'special'),
        ('local_other', 'Other Local Race', 'special'),
    ]
    for dt_id, dt_name, dt_cat in new_district_types:
        cur.execute("""
            INSERT INTO district_types (id, name, category)
            VALUES (%s, %s, %s)
            ON CONFLICT (id) DO NOTHING;
        """, (dt_id, dt_name, dt_cat))
    conn.commit()
    print(f"  Ensured {len(new_district_types)} district types exist")
    
    # ── Create jurisdiction ──
    print(f"\n--- Jurisdiction ---")
    cur.execute("""
        INSERT INTO jurisdictions (id, name, type, county, state)
        VALUES (%s, %s, %s, %s, %s)
        ON CONFLICT (id) DO UPDATE SET name = EXCLUDED.name;
    """, (JURISDICTION_ID, "Lake County", "county", "Lake", "IL"))
    conn.commit()
    print(f"  Upserted: {JURISDICTION_ID}")
    
    # ── Create elections (only new ones) ──
    print(f"\n--- Elections ({len(election_meta)}) ---")
    for eid, ename, edate, etype, eyear in election_meta:
        cur.execute("""
            INSERT INTO elections (id, name, date, type, year, state)
            VALUES (%s, %s, %s, %s, %s, 'IL')
            ON CONFLICT (id) DO UPDATE SET name = EXCLUDED.name;
        """, (eid, ename, edate, etype, eyear))
        print(f"  Upserted: {eid} ({ename})")
    conn.commit()
    
    # ── Clear existing Lake County results for idempotency ──
    print(f"\nClearing existing Lake County results...")
    cur.execute("DELETE FROM results WHERE source_file LIKE 'Lake County%';")
    print(f"  Deleted {cur.rowcount:,} results")
    cur.execute("DELETE FROM turnout WHERE source_file LIKE 'Lake County%';")
    print(f"  Deleted {cur.rowcount:,} turnout")
    conn.commit()
    
    # Reconnect after cleanup
    cur.close()
    conn.close()
    conn = psycopg2.connect(DB_URI, connect_timeout=30)
    conn.autocommit = False
    cur = conn.cursor()
    cur.execute("SET statement_timeout = 0;")
    
    # ── Load districts ──
    print(f"\n--- Districts ({len(all_d)}) ---")
    cur.execute("CREATE TEMP TABLE tmp_districts (LIKE districts INCLUDING ALL) ON COMMIT DROP;")
    n = copy_tuples(cur, "tmp_districts", ["id", "district_type_id", "number", "name"],
                    list(all_d.values()))
    cur.execute("""
        INSERT INTO districts (id, district_type_id, number, name)
        SELECT id, district_type_id, number, name FROM tmp_districts
        ON CONFLICT (id) DO UPDATE SET
            district_type_id = EXCLUDED.district_type_id,
            number = EXCLUDED.number,
            name = EXCLUDED.name;
    """)
    conn.commit()
    print(f"  Loaded {n}")
    
    # ── Load precincts ──
    print(f"\n--- Precincts ({len(all_p)}) ---")
    cur.execute("CREATE TEMP TABLE tmp_precincts (LIKE precincts INCLUDING ALL) ON COMMIT DROP;")
    precinct_tuples = list(all_p.values())
    n = copy_tuples(cur, "tmp_precincts",
                    ["id", "jurisdiction_id", "source_precinct_name", "ward", "precinct_number"],
                    precinct_tuples)
    cur.execute("""
        INSERT INTO precincts (id, jurisdiction_id, source_precinct_name, ward, precinct_number)
        SELECT id, jurisdiction_id, source_precinct_name, ward, precinct_number FROM tmp_precincts
        ON CONFLICT (id) DO UPDATE SET
            jurisdiction_id = EXCLUDED.jurisdiction_id,
            source_precinct_name = EXCLUDED.source_precinct_name,
            ward = EXCLUDED.ward,
            precinct_number = EXCLUDED.precinct_number;
    """)
    conn.commit()
    print(f"  Loaded {n}")
    
    # ── Load races ──
    print(f"\n--- Races ({len(all_r)}) ---")
    cur.execute("CREATE TEMP TABLE tmp_races (LIKE races INCLUDING ALL) ON COMMIT DROP;")
    n = copy_tuples(cur, "tmp_races",
                    ["id", "election_id", "district_id", "name", "source_contest_name", "race_type"],
                    list(all_r.values()))
    cur.execute("""
        INSERT INTO races (id, election_id, district_id, name, source_contest_name, race_type)
        SELECT id, election_id, district_id, name, source_contest_name, race_type FROM tmp_races
        ON CONFLICT (id) DO UPDATE SET
            election_id = EXCLUDED.election_id,
            district_id = EXCLUDED.district_id,
            name = EXCLUDED.name,
            source_contest_name = EXCLUDED.source_contest_name,
            race_type = EXCLUDED.race_type;
    """)
    conn.commit()
    print(f"  Loaded {n}")
    
    # ── Load candidates ──
    print(f"\n--- Candidates ({len(all_c)}) ---")
    cur.execute("CREATE TEMP TABLE tmp_cands (LIKE candidates INCLUDING ALL) ON COMMIT DROP;")
    n = copy_tuples(cur, "tmp_cands", ["id", "name", "party"], list(all_c.values()))
    cur.execute("""
        INSERT INTO candidates (id, name, party)
        SELECT id, name, party FROM tmp_cands
        ON CONFLICT (id) DO UPDATE SET name = EXCLUDED.name, party = EXCLUDED.party;
    """)
    conn.commit()
    print(f"  Loaded {n}")
    
    # ── Load precinct-district mappings ──
    valid_pids = set(p[0] for p in all_p.values())
    valid_dids = set(d[0] for d in all_d.values())
    pd_rows = [(pid, did, "Lake County") for (pid, did) in all_pd
               if pid in valid_pids and did in valid_dids]
    print(f"\n--- Precinct-Districts ({len(pd_rows)}) ---")
    cur.execute("CREATE TEMP TABLE tmp_pd (precinct_id text, district_id text, source_file text) ON COMMIT DROP;")
    n = copy_tuples(cur, "tmp_pd", ["precinct_id", "district_id", "source_file"], pd_rows)
    cur.execute("""
        INSERT INTO precinct_districts (precinct_id, district_id, source_file)
        SELECT t.precinct_id, t.district_id, t.source_file FROM tmp_pd t
        WHERE EXISTS (SELECT 1 FROM precincts p WHERE p.id = t.precinct_id)
          AND EXISTS (SELECT 1 FROM districts d WHERE d.id = t.district_id)
        ON CONFLICT (precinct_id, district_id) DO NOTHING;
    """)
    conn.commit()
    print(f"  Loaded {n}")
    
    # ── Deduplicate results (keep last occurrence for each unique key) ──
    print(f"\n--- Deduplicating results ---")
    result_dict = {}
    for r in all_results:
        key = (r[0], r[1], r[2], r[3])  # election_id, precinct_id, race_id, candidate_id
        result_dict[key] = r
    deduped_results = list(result_dict.values())
    print(f"  Before: {len(all_results):,}, After: {len(deduped_results):,} (removed {len(all_results) - len(deduped_results):,} duplicates)")
    
    # ── Load results (chunked) ──
    print(f"\n--- Results ({len(deduped_results):,}) ---")
    t1 = time.time()
    CHUNK = 200_000
    loaded = 0
    for i in range(0, len(deduped_results), CHUNK):
        chunk = deduped_results[i:i+CHUNK]
        cur.execute("""CREATE TEMP TABLE tmp_results (
            election_id text, precinct_id text, race_id text, candidate_id text,
            votes integer, source_file text, source_contest_name text,
            source_precinct_name text, source_candidate_name text
        ) ON COMMIT DROP;""")
        n = copy_tuples(cur, "tmp_results",
                        ["election_id", "precinct_id", "race_id", "candidate_id", "votes",
                         "source_file", "source_contest_name", "source_precinct_name",
                         "source_candidate_name"],
                        chunk)
        cur.execute("""
            INSERT INTO results (election_id, precinct_id, race_id, candidate_id, votes,
                                 source_file, source_contest_name, source_precinct_name,
                                 source_candidate_name)
            SELECT election_id, precinct_id, race_id, candidate_id, votes,
                   source_file, source_contest_name, source_precinct_name,
                   source_candidate_name
            FROM tmp_results
            ON CONFLICT (election_id, race_id, precinct_id, candidate_id)
            DO UPDATE SET votes = EXCLUDED.votes,
                          source_file = EXCLUDED.source_file,
                          source_contest_name = EXCLUDED.source_contest_name,
                          source_precinct_name = EXCLUDED.source_precinct_name,
                          source_candidate_name = EXCLUDED.source_candidate_name;
        """)
        conn.commit()
        loaded += n
        elapsed = time.time() - t1
        print(f"  Chunk {i//CHUNK + 1}: {loaded:,} / {len(deduped_results):,}  ({elapsed:.1f}s)")
    
    print(f"  Total loaded: {loaded:,} in {time.time()-t1:.1f}s")
    
    # ── Load turnout ──
    print(f"\n--- Turnout ({len(all_turnout):,}) ---")
    t1 = time.time()
    cur.execute("""CREATE TEMP TABLE tmp_turnout (
        election_id text, precinct_id text, registered_voters integer,
        ballots_cast integer, turnout_pct numeric, source_file text,
        source_precinct_name text
    ) ON COMMIT DROP;""")
    n = copy_tuples(cur, "tmp_turnout",
                    ["election_id", "precinct_id", "registered_voters", "ballots_cast",
                     "turnout_pct", "source_file", "source_precinct_name"],
                    all_turnout)
    cur.execute("""
        INSERT INTO turnout (election_id, precinct_id, registered_voters, ballots_cast,
                             turnout_pct, source_file, source_precinct_name)
        SELECT election_id, precinct_id, registered_voters, ballots_cast,
               turnout_pct, source_file, source_precinct_name
        FROM tmp_turnout
        ON CONFLICT (election_id, precinct_id)
        DO UPDATE SET registered_voters = EXCLUDED.registered_voters,
                      ballots_cast = EXCLUDED.ballots_cast,
                      turnout_pct = EXCLUDED.turnout_pct,
                      source_file = EXCLUDED.source_file,
                      source_precinct_name = EXCLUDED.source_precinct_name;
    """)
    conn.commit()
    print(f"  Loaded {n:,} in {time.time()-t1:.1f}s")
    
    # ── Refresh materialized view ──
    print("\nRefreshing materialized view...")
    cur.execute("REFRESH MATERIALIZED VIEW mv_race_precinct_results;")
    conn.commit()
    
    print("Running ANALYZE...")
    cur.execute("ANALYZE;")
    conn.commit()
    
    # ── Verification ──
    print("\n--- Verification ---")
    cur.execute("SELECT COUNT(*) FROM results WHERE source_file LIKE 'Lake County%';")
    lake_results = cur.fetchone()[0]
    print(f"Lake County results in DB: {lake_results:,}")
    
    cur.execute("SELECT source_file, COUNT(*) FROM results WHERE source_file LIKE 'Lake County%' GROUP BY source_file ORDER BY source_file;")
    for row in cur.fetchall():
        print(f"  {row[0]}: {row[1]:,}")
    
    cur.execute("SELECT COUNT(*) FROM precincts WHERE jurisdiction_id = %s;", (JURISDICTION_ID,))
    print(f"\nLake County precincts: {cur.fetchone()[0]:,}")
    
    cur.execute("SELECT COUNT(*) FROM turnout WHERE source_file LIKE 'Lake County%';")
    print(f"Lake County turnout rows: {cur.fetchone()[0]:,}")
    
    cur.execute("SELECT COUNT(*) FROM results;")
    print(f"\nTotal results in DB: {cur.fetchone()[0]:,}")
    
    cur.execute("SELECT COUNT(*) FROM precincts;")
    print(f"Total precincts in DB: {cur.fetchone()[0]:,}")
    
    cur.close()
    conn.close()
    
    total_time = time.time() - t0
    print(f"\n{'='*60}")
    print(f"LAKE COUNTY LOAD COMPLETE in {total_time:.1f}s")
    print(f"{'='*60}")


if __name__ == "__main__":
    main()
