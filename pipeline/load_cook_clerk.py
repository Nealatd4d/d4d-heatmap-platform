#!/usr/bin/env python3
"""
Load precinct-level election data from Cook County Clerk results website.
Downloads race-by-race XLSX files, parses precinct sheets, and uploads to Supabase.

Supports:
  - 2025 Consolidated (results425.cookcountyclerkil.gov)
  - 2026 Primary (results326.cookcountyclerkil.gov)

Schema mapping (verified against actual Supabase tables):
  elections:   id, name, date, type, year, state
  races:       id, election_id, district_id, name, source_contest_name, race_type
  candidates:  id, name, party  (NO race_id — linked via results)
  results:     election_id, precinct_id, race_id, candidate_id, votes, vote_percentage,
               source_file, source_contest_name, source_precinct_name, source_candidate_name
  turnout:     election_id, precinct_id, registered_voters, ballots_cast, turnout_pct,
               source_file, source_precinct_name
"""

import os, sys, json, hashlib, time, re, traceback
import requests
import openpyxl
from io import BytesIO
from collections import defaultdict

# ── Supabase config ──
SUPA_URL = "https://nfjfqdffulhqhszhlymo.supabase.co"
SUPA_KEY = "eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9.eyJpc3MiOiJzdXBhYmFzZSIsInJlZiI6Im5mamZxZGZmdWxocWhzemhseW1vIiwicm9sZSI6InNlcnZpY2Vfcm9sZSIsImlhdCI6MTc3Mzk1MDMwOCwiZXhwIjoyMDg5NTI2MzA4fQ.zy7Bfl5cOZGF_CU2yiHn3sd24olsx_Hzc985Ov_t5Ys"
HEADERS = {
    "apikey": SUPA_KEY,
    "Authorization": f"Bearer {SUPA_KEY}",
    "Content-Type": "application/json",
    "Prefer": "return=minimal,resolution=merge-duplicates"
}

# ── Election configs ──
ELECTIONS = {
    "2025_consolidated": {
        "base_url": "https://results425.cookcountyclerkil.gov",
        "name": "2025 Consolidated Election",
        "date": "2025-04-01",
        "id": "2025_consolidated",
        "type": "consolidated",
        "year": 2025,
        "state": "IL",
        "source_file": "Cook County Clerk 2025 Consolidated",
        "max_contest_id": 1000,
    },
    "2026_primary": {
        "base_url": "https://results326.cookcountyclerkil.gov",
        "name": "2026 Gubernatorial Primary",
        "date": "2026-03-17",
        "id": "2026_primary",
        "type": "primary",
        "year": 2026,
        "state": "IL",
        "source_file": "Cook County Clerk 2026 Primary",
        "max_contest_id": 1000,
    },
}

# ── Township precinct ID prefix mapping ──
TOWNSHIP_CODES = {
    "barrington": "70",
    "berwyn": "99",
    "bloom": "71",
    "bremen": "72",
    "calumet": "73",
    "cicero": "98",
    "elk grove": "74",
    "evanston": "75",
    "hanover": "76",
    "lemont": "77",
    "leyden": "78",
    "lyons": "79",
    "maine": "80",
    "new trier": "81",
    "niles": "82",
    "northfield": "83",
    "norwood park": "84",
    "oak park": "85",
    "orland": "86",
    "palatine": "87",
    "palos": "88",
    "proviso": "89",
    "rich": "90",
    "river forest": "91",
    "riverside": "92",
    "schaumburg": "93",
    "stickney": "94",
    "thornton": "95",
    "wheeling": "96",
    "worth": "97",
}


def make_precinct_id(jurisdiction, source_name):
    """MD5-based precinct ID matching our Supabase schema."""
    raw = f"{jurisdiction}|{source_name}".lower().strip()
    return hashlib.md5(raw.encode()).hexdigest()[:16]


def parse_precinct_name(precinct_str):
    """
    Parse precinct string from Cook County Clerk XLSX.
    Input:  "Wheeling Precinct 2" or "Elk Grove Precinct 14" or "Evanston Ward 1 Precinct 3"
    Returns: (township_name, precinct_number, precinctid_for_geojson)
    """
    s = precinct_str.strip()

    # Handle Evanston ward-precinct format FIRST (before generic pattern)
    m = re.match(r'^Evanston\s+Ward\s+(\d+)\s+Precinct\s+(\d+)$', s, re.IGNORECASE)
    if m:
        ward = int(m.group(1))
        precinct = int(m.group(2))
        precinctid = f"75{ward:02d}{precinct:03d}"
        return "evanston", f"W{ward}P{precinct}", precinctid

    # Handle Berwyn ward-precinct format
    m = re.match(r'^Berwyn\s+Ward\s+(\d+)\s+Precinct\s+(\d+)$', s, re.IGNORECASE)
    if m:
        ward = int(m.group(1))
        precinct = int(m.group(2))
        precinctid = f"99{ward:02d}{precinct:03d}"
        return "berwyn", f"W{ward}P{precinct}", precinctid

    # Handle "Township Precinct N" format
    m = re.match(r'^(.+?)\s+Precinct\s+(\d+)$', s, re.IGNORECASE)
    if m:
        township = m.group(1).strip().lower()
        precinct_num = int(m.group(2))
        code = TOWNSHIP_CODES.get(township)
        if code:
            precinctid = f"{code}{precinct_num:05d}"
            return township, precinct_num, precinctid
        else:
            return township, precinct_num, None

    return None, None, None


def classify_race(race_name):
    """
    Classify a race name into our race_type taxonomy.
    Returns (race_type, simplified_name).
    """
    name = race_name.strip()
    lower = name.lower()

    # Township offices
    if 'township' in lower and ('supervisor' in lower or 'assessor' in lower or
                                  'highway commissioner' in lower or 'collector' in lower):
        return 'township_office', name
    if re.match(r'.*\btownship\b.*\b(clerk|trustee)\b', lower):
        return 'township_office', name
    if 'township' in lower and ('pension reform' in lower or 'unfunded mandates' in lower or
                                  'redistricting reform' in lower or 'mental health' in lower):
        return 'referendum', name

    # Municipal offices
    if any(t in lower for t in ['president, village', 'mayor,', 'clerk, village', 'clerk, city',
                                  'trustee, village', 'trustee, city', 'treasurer, village',
                                  'treasurer, city', 'alderman', 'commissioner, village',
                                  'commissioner, city']):
        return 'municipal', name

    # Library
    if 'library' in lower:
        return 'library', name

    # Park district
    if 'park' in lower and ('commissioner' in lower or 'trustee' in lower or
                             'president' in lower or 'board' in lower):
        return 'park_district', name

    # School district / school board
    if any(t in lower for t in ['school b.m.', 'school board', 'school district',
                                  'school trustee', 'hsd ', 'elementary', 'high school',
                                  'community college', 'unit school']):
        return 'school_district', name

    # Fire district
    if 'fire' in lower:
        return 'fire_district', name

    # Sanitary / water district
    if any(t in lower for t in ['water reclamation', 'sanitary', 'water district']):
        return 'water_reclamation', name

    # Judge / judicial
    if any(t in lower for t in ['judge', 'justice', 'judicial', 'circuit court', 'appellate']):
        return 'judge', name

    # Referendum
    if any(t in lower for t in ['referendum', 'issue bonds', 'increase limiting rate',
                                  'parks and recreation', 'proposition', 'advisory']):
        return 'referendum', name

    # Municipal ward races
    if 'ward' in lower and ('alderman' in lower or 'council' in lower):
        return 'municipal_ward', name

    # Congress / U.S. Representative
    if 'congress' in lower or 'representative in congress' in lower or 'u.s. representative' in lower:
        return 'congress', name

    # U.S. Senate
    if 'senator, u.s.' in lower:
        return 'us_senate', name

    # State senate
    if 'state senator' in lower or ('senate' in lower and 'u.s.' not in lower):
        return 'state_senate', name

    # State house / State Representative
    if 'state representative' in lower:
        return 'state_rep', name
    if 'representative' in lower and 'congress' not in lower and 'u.s.' not in lower:
        return 'state_rep', name

    # Governor
    if 'governor' in lower:
        return 'governor', name

    # Attorney General
    if 'attorney general' in lower:
        return 'attorney_general', name

    # Secretary of State
    if 'secretary of state' in lower:
        return 'secretary_of_state', name

    # Comptroller
    if 'comptroller' in lower:
        return 'comptroller', name

    # State Treasurer
    if 'treasurer, state' in lower or 'treasurer, state of' in lower:
        return 'treasurer_state', name

    # Cook County offices
    if 'assessor, cook county' in lower:
        return 'cook_assessor', name
    if 'clerk, cook county' in lower:
        return 'cook_clerk', name
    if 'treasurer, cook county' in lower:
        return 'cook_treasurer', name
    if 'sheriff, cook county' in lower:
        return 'cook_sheriff', name
    if 'board president, cook county' in lower:
        return 'cook_president', name
    if 'board of review' in lower:
        return 'board_of_review', name

    # State Central Committeeperson
    if 'committeeperson, dem state central' in lower or 'committeeperson, rep state central' in lower:
        return 'state_central_comm', name

    # Township Committeeperson (party)
    if re.match(r'(dem|rep|lib) committeeperson', lower):
        return 'township_committeeperson', name

    # County commissioner
    if ('commissioner' in lower and 'county board' in lower) or ('commissioner' in lower and 'cook county' in lower):
        return 'cook_commissioner', name

    # Alderperson (suburban variant of alderman)
    if 'alderperson' in lower:
        return 'municipal_ward', name

    # Council member
    if 'council member' in lower or 'councilmember' in lower:
        return 'municipal_ward', name

    # Township clerk/trustee (not caught by earlier rules)
    if re.match(r'^(clerk|trustee),.*township', lower):
        return 'township_office', name

    # Park district commissioner (Cmsnr.)
    if 'cmsnr.' in lower and 'park' in lower:
        return 'park_district', name

    # Fire protection district
    if 'fpd' in lower or 'fire protection' in lower:
        return 'fire_district', name

    # Community college trustee
    if 'college' in lower and 'trustee' in lower:
        return 'school_district', name

    # Trustee of Schools
    if 'trustee of schools' in lower:
        return 'school_district', name

    return 'other', name


def download_contest(base_url, contest_id, session):
    """Download a single contest XLSX and return workbook or None."""
    url = f"{base_url}/Home/SummaryExport?contestId={contest_id}"
    try:
        r = session.get(url, timeout=30)
        if r.status_code != 200:
            return None
        # Check if it's actually an XLSX (some return HTML error pages)
        content_type = r.headers.get('Content-Type', '')
        if 'html' in content_type and len(r.content) < 1000:
            return None
        wb = openpyxl.load_workbook(BytesIO(r.content), read_only=True)
        return wb
    except Exception as e:
        return None


def parse_contest_xlsx(wb):
    """
    Parse a contest XLSX workbook.
    Returns: {
        'race_name': str,
        'vote_for': int,
        'candidates': [{'name': str, 'party': str}],
        'precincts': [{'precinct_name': str, 'registered': int, 'ballots': int,
                        'candidate_votes': {name: int}, 'total_votes': int}],
    }
    """
    result = {
        'race_name': '',
        'vote_for': 1,
        'candidates': [],
        'precincts': [],
    }

    # Parse Summary sheet for race name and candidates
    if 'Summary' in wb.sheetnames:
        ws = wb['Summary']
        rows = list(ws.iter_rows(values_only=True))
        if len(rows) >= 3:
            race_info = rows[1]
            result['race_name'] = str(race_info[0] or '').strip()
            vote_for_str = str(race_info[1] or '')
            m = re.search(r'Vote [Ff]or (\d+)', vote_for_str)
            if m:
                result['vote_for'] = int(m.group(1))

            for row in rows[3:]:
                if row[0] and str(row[0]).strip():
                    cand_name = str(row[0]).strip()
                    # Skip total/header rows
                    if cand_name.lower() in ('total', 'totals', ''):
                        continue
                    result['candidates'].append({
                        'name': cand_name,
                        'party': str(row[1] or '').strip()
                    })

    # Parse Precinct sheet for precinct-level data
    if 'Precinct' in wb.sheetnames:
        ws = wb['Precinct']
        rows = list(ws.iter_rows(values_only=True))
        if len(rows) >= 2:
            header = rows[0]
            # Header: Precinct, Registered Voters, Ballots Cast, [candidate names...], Total Votes
            cand_names = [str(h).strip() for h in header[3:-1] if h]

            for row in rows[1:]:
                precinct_name = str(row[0] or '').strip()
                if not precinct_name or precinct_name.lower() in ('suburban cook county', 'total', 'totals', ''):
                    continue

                registered = _parse_int(row[1])
                ballots = _parse_int(row[2])
                total_votes = _parse_int(row[-1])

                cand_votes = {}
                for i, cname in enumerate(cand_names):
                    val = _parse_int(row[3 + i]) if (3 + i) < len(row) else 0
                    cand_votes[cname] = val

                result['precincts'].append({
                    'precinct_name': precinct_name,
                    'registered': registered,
                    'ballots': ballots,
                    'candidate_votes': cand_votes,
                    'total_votes': total_votes
                })

    return result


def _parse_int(v):
    """Parse a value that might be int, string with commas, or empty."""
    if v is None:
        return 0
    s = str(v).replace(',', '').strip()
    try:
        return int(float(s))
    except (ValueError, TypeError):
        return 0


def supa_upsert(table, data):
    """Upsert rows into Supabase via PostgREST."""
    if not data:
        return None
    r = requests.post(
        f"{SUPA_URL}/rest/v1/{table}",
        headers=HEADERS,
        json=data
    )
    if r.status_code not in (200, 201):
        print(f"  !! Supabase error on {table}: {r.status_code} {r.text[:300]}")
    return r


def ensure_election(cfg):
    """Ensure election record exists in Supabase."""
    data = {
        "id": cfg["id"],
        "name": cfg["name"],
        "date": cfg["date"],
        "type": cfg["type"],
        "year": cfg["year"],
        "state": cfg["state"],
    }
    r = supa_upsert("elections", [data])
    if r and r.status_code in (200, 201):
        print(f"  Election '{cfg['name']}' ensured in Supabase.")
    else:
        print(f"  !! Failed to create election record. Cannot continue.")
        sys.exit(1)


def process_election(election_key):
    """Download and process all contests for an election."""
    cfg = ELECTIONS[election_key]
    base_url = cfg["base_url"]
    election_id = cfg["id"]
    source_file = cfg["source_file"]
    max_id = cfg["max_contest_id"]

    print(f"\n{'='*60}")
    print(f"Processing: {cfg['name']}")
    print(f"Source: {base_url}")
    print(f"{'='*60}")

    # Ensure election exists
    ensure_election(cfg)

    # Create HTTP session with browser UA
    session = requests.Session()
    session.headers.update({
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/131.0.0.0 Safari/537.36"
    })

    stats = {
        "contests_downloaded": 0,
        "contests_empty": 0,
        "contests_error": 0,
        "total_races": 0,
        "total_precinct_rows": 0,
        "total_results_uploaded": 0,
        "total_turnout_uploaded": 0,
        "unmapped_precincts": set(),
        "race_types": defaultdict(int),
    }

    # Batch accumulators
    race_batch = []
    candidate_batch = []
    result_batch = []
    turnout_batch = []
    turnout_seen = set()  # (election_id, precinct_id) to avoid duplicates
    candidate_seen = set()  # deduplicate candidates across contests

    BATCH_SIZE = 500

    def flush_batches(force=False):
        nonlocal race_batch, candidate_batch, result_batch, turnout_batch
        if race_batch:
            supa_upsert("races", race_batch)
            stats["total_races"] += len(race_batch)
            race_batch = []
        if candidate_batch:
            supa_upsert("candidates", candidate_batch)
            candidate_batch = []
        if result_batch:
            # Upload results in chunks of 500 (Supabase limit)
            for i in range(0, len(result_batch), 500):
                chunk = result_batch[i:i+500]
                r = supa_upsert("results", chunk)
                if r and r.status_code in (200, 201):
                    stats["total_results_uploaded"] += len(chunk)
            result_batch = []
        if turnout_batch:
            r = supa_upsert("turnout", turnout_batch)
            if r and r.status_code in (200, 201):
                stats["total_turnout_uploaded"] += len(turnout_batch)
            turnout_batch = []

    for contest_id in range(1, max_id + 1):
        if contest_id % 50 == 0:
            print(f"  Progress: contest {contest_id}/{max_id} | "
                  f"races={stats['total_races']} results={stats['total_results_uploaded']}")
            flush_batches()

        wb = download_contest(base_url, contest_id, session)
        if wb is None:
            stats["contests_error"] += 1
            continue

        if 'Precinct' not in wb.sheetnames:
            stats["contests_empty"] += 1
            wb.close()
            continue

        parsed = parse_contest_xlsx(wb)
        wb.close()

        if not parsed['race_name'] or not parsed['precincts']:
            stats["contests_empty"] += 1
            continue

        stats["contests_downloaded"] += 1

        # Classify the race
        race_type, race_display = classify_race(parsed['race_name'])
        stats["race_types"][race_type] += 1

        # Generate race ID
        race_id = hashlib.md5(f"{election_id}|{parsed['race_name']}".encode()).hexdigest()[:16]

        # -- races table: id, election_id, district_id, name, source_contest_name, race_type --
        race_batch.append({
            "id": race_id,
            "election_id": election_id,
            "district_id": None,
            "name": parsed['race_name'],
            "source_contest_name": parsed['race_name'],
            "race_type": race_type,
        })

        # -- candidates table: id, name, party (NO race_id) --
        for cand in parsed['candidates']:
            cand_id = hashlib.md5(f"{race_id}|{cand['name']}".encode()).hexdigest()[:16]
            if cand_id not in candidate_seen:
                candidate_seen.add(cand_id)
                candidate_batch.append({
                    "id": cand_id,
                    "name": cand['name'],
                    "party": cand['party'] if cand['party'] else None,
                })

        # Process precinct results
        for p in parsed['precincts']:
            township, precinct_num, precinctid = parse_precinct_name(p['precinct_name'])

            if precinctid is None:
                stats["unmapped_precincts"].add(p['precinct_name'])
                continue

            # Generate our MD5 precinct_id
            pid = make_precinct_id("cook_suburban", precinctid)

            # -- turnout table: election_id, precinct_id, registered_voters, ballots_cast,
            #                   turnout_pct, source_file, source_precinct_name --
            tk = (election_id, pid)
            if tk not in turnout_seen:
                turnout_seen.add(tk)
                turnout_pct = (p['ballots'] / p['registered'] * 100) if p['registered'] > 0 else 0
                turnout_batch.append({
                    "election_id": election_id,
                    "precinct_id": pid,
                    "registered_voters": p['registered'],
                    "ballots_cast": p['ballots'],
                    "turnout_pct": round(turnout_pct, 2),
                    "source_file": source_file,
                    "source_precinct_name": p['precinct_name'],
                })

            # -- results table: election_id, precinct_id, race_id, candidate_id, votes,
            #                   vote_percentage, source_file, source_contest_name,
            #                   source_precinct_name, source_candidate_name --
            for cand_name, votes in p['candidate_votes'].items():
                cand_id = hashlib.md5(f"{race_id}|{cand_name}".encode()).hexdigest()[:16]
                total = p['total_votes'] if p['total_votes'] > 0 else 1
                result_batch.append({
                    "election_id": election_id,
                    "precinct_id": pid,
                    "race_id": race_id,
                    "candidate_id": cand_id,
                    "votes": votes,
                    "vote_percentage": round(votes / total * 100, 2) if total > 0 else 0,
                    "source_file": source_file,
                    "source_contest_name": parsed['race_name'],
                    "source_precinct_name": p['precinct_name'],
                    "source_candidate_name": cand_name,
                })

            stats["total_precinct_rows"] += 1

        # Flush if batch is large
        if len(result_batch) >= BATCH_SIZE:
            flush_batches()

        # Rate limiting
        time.sleep(0.05)

    # Final flush
    flush_batches(force=True)

    print(f"\n{'='*60}")
    print(f"COMPLETE: {cfg['name']}")
    print(f"  Contests downloaded: {stats['contests_downloaded']}")
    print(f"  Contests empty/skipped: {stats['contests_empty']}")
    print(f"  Contests error: {stats['contests_error']}")
    print(f"  Total races: {stats['total_races']}")
    print(f"  Total precinct rows parsed: {stats['total_precinct_rows']}")
    print(f"  Total results uploaded: {stats['total_results_uploaded']}")
    print(f"  Total turnout rows uploaded: {stats['total_turnout_uploaded']}")
    print(f"\n  Race types:")
    for rt, count in sorted(stats["race_types"].items(), key=lambda x: -x[1]):
        print(f"    {rt}: {count}")
    if stats["unmapped_precincts"]:
        print(f"\n  Unmapped precincts ({len(stats['unmapped_precincts'])}):")
        for p in sorted(stats["unmapped_precincts"])[:20]:
            print(f"    - {p}")
        if len(stats["unmapped_precincts"]) > 20:
            print(f"    ... and {len(stats['unmapped_precincts']) - 20} more")
    print(f"{'='*60}")

    return stats


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Usage: python load_cook_clerk.py <election_key>")
        print("  election_key: 2025_consolidated | 2026_primary | all")
        sys.exit(1)

    key = sys.argv[1]

    if key == "all":
        for k in ELECTIONS:
            process_election(k)
    elif key in ELECTIONS:
        process_election(key)
    else:
        print(f"Unknown election key: {key}")
        print(f"Available: {', '.join(ELECTIONS.keys())}")
        sys.exit(1)
