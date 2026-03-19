#!/usr/bin/env python3
"""
D4D Election Heatmap Platform — Data Pipeline
Normalizes Chicago SBOE CSVs + Cook County Clerk XLSX into a unified election database.
Output: data/elections.json — precinct-level results for all elections and contests.
"""

import csv
import json
import re
import os
import sys
from collections import defaultdict

# ── Configuration ──────────────────────────────────────────────

WORKSPACE = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
DATA_DIR = os.path.join(WORKSPACE, 'data')
ROOT = os.path.dirname(WORKSPACE)  # /home/user/workspace

# Chicago SBOE CSV files
CHICAGO_CSVS = {
    # '2022_primary': os.path.join(ROOT, 'chicago_cook_2022GP.csv'),  # Old boundaries, skip
    '2022_general': os.path.join(ROOT, 'chicago_cook_2022GE.csv'),
    '2024_primary': os.path.join(ROOT, 'chicago_cook_2024GP.csv'),
    '2024_general': os.path.join(ROOT, 'chicago_cook_2024GE.csv'),
}

# Cook County Clerk XLSX files (suburban Cook)
COOK_XLSX = {
    '2022_primary': os.path.join(ROOT, '62822-6.xlsx'),
    '2022_general': os.path.join(ROOT, '112822.xlsx'),
    '2024_primary': os.path.join(ROOT, '31924-10.xlsx'),
    '2024_general': os.path.join(ROOT, '11524-11.xlsx'),
}

# Contest name normalization patterns
CONTEST_PATTERNS = {
    'state_rep': [
        (r'(\d+)(?:ST|ND|RD|TH)\s+REPRESENTATIVE', 'state_rep_{0}'),
        (r'State Representative,?\s*(\d+)', 'state_rep_{0}'),
    ],
    'state_senate': [
        (r'(\d+)(?:ST|ND|RD|TH)\s+SENATE', 'state_senate_{0}'),
        (r'State Senator,?\s*(\d+)', 'state_senate_{0}'),
    ],
    'congress': [
        (r'(\d+)(?:ST|ND|RD|TH)\s+CONGRESS(?:IONAL)?', 'congress_{0}'),
        (r'U\.?S\.?\s+Representative,?\s*(\d+)', 'congress_{0}'),
    ],
    'us_senate': [
        (r'(?:UNITED STATES|U\.?S\.?)\s+SENATOR', 'us_senate'),
        (r'Senator,?\s*U\.?S\.?', 'us_senate'),
    ],
    'governor': [
        (r'GOVERNOR', 'governor'),
    ],
}


def normalize_precinct_id(precinct_name):
    """Convert 'Ward 01 Precinct 05' → '01005' (5-digit ward-precinct ID)"""
    m = re.match(r'Ward\s+(\d+)\s+Precinct\s+(\d+)', precinct_name, re.IGNORECASE)
    if m:
        ward = int(m.group(1))
        prec = int(m.group(2))
        return f"{ward:02d}{prec:03d}"
    return None


def normalize_contest_name(raw_name):
    """
    Normalize contest names to canonical keys.
    Returns (category, key, district_num_or_none)
    e.g. ('state_rep', 'state_rep_13', 13) or ('us_senate', 'us_senate', None)
    """
    raw = raw_name.strip().upper()
    
    for category, patterns in CONTEST_PATTERNS.items():
        for pattern, key_tmpl in patterns:
            m = re.search(pattern, raw, re.IGNORECASE)
            if m:
                if m.groups():
                    district = int(m.group(1))
                    key = key_tmpl.format(district)
                    return (category, key, district)
                else:
                    return (category, key_tmpl, None)
    
    return (None, None, None)


def parse_chicago_csv(filepath, election_key):
    """
    Parse a Chicago SBOE CSV into normalized precinct data.
    Returns dict: precinct_id → {registration, ballots, contests: {contest_key: {candidates: [...]}}}
    """
    print(f"  Parsing {os.path.basename(filepath)}...")
    
    precincts = {}
    
    with open(filepath, 'r', errors='ignore') as f:
        reader = csv.DictReader(f)
        for row in reader:
            precinct_name = row.get('PrecinctName', '').strip()
            if not precinct_name:
                continue
            
            pid = normalize_precinct_id(precinct_name)
            if not pid:
                continue
            
            contest_name = row.get('ContestName', '').strip()
            candidate_name = row.get('CandidateName', '').strip()
            votes = int(row.get('VoteCount', '0').strip() or '0')
            registration = int(row.get('Registration', '0').strip() or '0')
            party = row.get('PartyName', '').strip()
            
            if pid not in precincts:
                precincts[pid] = {
                    'precinct_id': pid,
                    'precinct_name': precinct_name,
                    'jurisdiction': 'chicago',
                    'registration': 0,
                    'ballots': 0,
                    'contests': {}
                }
            
            p = precincts[pid]
            
            # Blank contest = registration/ballot summary row
            if not contest_name:
                p['registration'] = registration
                p['ballots'] = votes
                continue
            
            # Skip under/over votes and write-ins for our purposes
            if candidate_name.upper() in ('UNDER VOTES', 'OVER VOTES', 'WRITE-IN', 'NO CANDIDATE', ''):
                continue
            
            # Normalize contest name
            category, contest_key, district = normalize_contest_name(contest_name)
            
            if not contest_key:
                # Store raw contest name for non-matched contests  
                contest_key = re.sub(r'[^a-z0-9]+', '_', contest_name.lower()).strip('_')
                category = 'other'
            
            if contest_key not in p['contests']:
                p['contests'][contest_key] = {
                    'raw_name': contest_name,
                    'category': category,
                    'district': district,
                    'party': party,
                    'candidates': []
                }
            
            p['contests'][contest_key]['candidates'].append({
                'name': candidate_name,
                'votes': votes,
                'party': party
            })
    
    print(f"    → {len(precincts)} precincts parsed")
    return precincts


def parse_cook_xlsx(filepath, election_key):
    """
    Parse a Cook County Clerk XLSX into normalized precinct data.
    These use a different format: multi-sheet workbook with one sheet per race.
    Returns dict: precinct_id → {registration, ballots, contests: {...}}
    """
    try:
        import openpyxl
    except ImportError:
        print(f"  ⚠ openpyxl not available, skipping {os.path.basename(filepath)}")
        return {}
    
    print(f"  Parsing {os.path.basename(filepath)}...")
    
    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    precincts = {}
    
    # First sheet is document map - tells us what's on each sheet
    doc_map_sheet = wb[wb.sheetnames[0]]
    sheet_map = {}
    for row in doc_map_sheet.iter_rows(min_row=1, values_only=True):
        if row and row[0] and isinstance(row[0], str):
            # Format: "SheetN" or just descriptions
            pass
    
    # Process each sheet (skip the document map)
    for sheet_idx, sheet_name in enumerate(wb.sheetnames):
        if sheet_idx == 0:
            continue  # Skip document map
        
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        
        if not rows:
            continue
        
        # Find the contest name from the header area
        contest_name = None
        header_row_idx = None
        
        for i, row in enumerate(rows[:10]):
            if row and row[0]:
                val = str(row[0]).strip()
                # Look for the header row with "Jurisdiction" or precinct data
                if val.upper() in ('JURISDICTION', 'PRECINCT'):
                    header_row_idx = i
                    break
                # Contest name is usually in first few rows
                if not contest_name and val and not val.startswith(('Psd', 'PrecinctCanvass', 'District')):
                    # Check if this looks like a contest description
                    if any(kw in val.upper() for kw in ('REPRESENTATIVE', 'SENATOR', 'CONGRESS', 'JUDGE', 'GOVERNOR', 'COMMISSIONER', 'PRESIDENT')):
                        contest_name = val
        
        if not header_row_idx or not contest_name:
            continue
        
        # Parse header to find candidate columns
        header = rows[header_row_idx]
        # ... This gets complex with the Cook County format
        # For now, we'll focus on Chicago data and add Cook County parsing later
    
    wb.close()
    print(f"    → {len(precincts)} precincts parsed (Cook County)")
    return precincts


def build_election_database():
    """Build the unified election database from all sources."""
    
    print("=" * 60)
    print("D4D Election Heatmap Platform — Data Pipeline")
    print("=" * 60)
    
    database = {
        'meta': {
            'generated': None,
            'elections': {},
            'precinct_count': 0,
        },
        'elections': {}
    }
    
    # ── Parse Chicago SBOE CSVs ──
    print("\n── Chicago SBOE CSVs ──")
    for election_key, filepath in CHICAGO_CSVS.items():
        if not os.path.exists(filepath):
            print(f"  ⚠ Missing: {filepath}")
            continue
        
        precincts = parse_chicago_csv(filepath, election_key)
        
        # Organize by contest for easier lookup
        election_data = {
            'key': election_key,
            'jurisdiction': 'chicago',
            'precinct_count': len(precincts),
            'precincts': {}
        }
        
        # Store precinct data
        for pid, pdata in precincts.items():
            turnout = round(pdata['ballots'] / pdata['registration'] * 100, 1) if pdata['registration'] > 0 else 0
            
            election_data['precincts'][pid] = {
                'name': pdata['precinct_name'],
                'reg': pdata['registration'],
                'bal': pdata['ballots'],
                'to': turnout,
                'races': {}
            }
            
            # Flatten contests
            for ckey, cdata in pdata['contests'].items():
                # Sort candidates by votes desc
                cands = sorted(cdata['candidates'], key=lambda x: x['votes'], reverse=True)
                total_votes = sum(c['votes'] for c in cands)
                
                race = {
                    'raw': cdata['raw_name'],
                    'cat': cdata['category'],
                    'dist': cdata['district'],
                    'party': cdata['party'],
                    'total': total_votes,
                    'cands': []
                }
                
                for c in cands:
                    pct = round(c['votes'] / total_votes * 100, 1) if total_votes > 0 else 0
                    race['cands'].append({
                        'n': c['name'],
                        'v': c['votes'],
                        'p': pct
                    })
                
                election_data['precincts'][pid]['races'][ckey] = race
        
        database['elections'][election_key] = election_data
        
        # Summarize contests found
        all_contests = set()
        for pid, pdata in precincts.items():
            for ckey in pdata['contests']:
                all_contests.add(ckey)
        
        state_rep = [c for c in all_contests if c.startswith('state_rep_')]
        state_sen = [c for c in all_contests if c.startswith('state_senate_')]
        congress = [c for c in all_contests if c.startswith('congress_')]
        
        print(f"    Contests: {len(all_contests)} total, {len(state_rep)} state rep, {len(state_sen)} state senate, {len(congress)} congressional")
    
    # ── Summary ──
    total_precincts = set()
    for ek, ed in database['elections'].items():
        for pid in ed['precincts']:
            total_precincts.add(pid)
    
    database['meta']['precinct_count'] = len(total_precincts)
    database['meta']['elections'] = {
        k: {
            'precinct_count': v['precinct_count'],
            'jurisdiction': v['jurisdiction']
        } for k, v in database['elections'].items()
    }
    
    import datetime
    database['meta']['generated'] = datetime.datetime.now().isoformat()
    
    print(f"\n{'=' * 60}")
    print(f"Total unique precincts: {len(total_precincts)}")
    print(f"Elections: {list(database['elections'].keys())}")
    
    return database


def save_database(database, output_path):
    """Save the election database to JSON."""
    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    
    with open(output_path, 'w') as f:
        json.dump(database, f, separators=(',', ':'))
    
    size_mb = os.path.getsize(output_path) / (1024 * 1024)
    print(f"\nSaved: {output_path} ({size_mb:.1f} MB)")


if __name__ == '__main__':
    database = build_election_database()
    output_path = os.path.join(DATA_DIR, 'elections.json')
    save_database(database, output_path)
    print("\nDone.")
